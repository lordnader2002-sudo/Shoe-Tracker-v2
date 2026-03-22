"""
Excel export module for sneaker release reports.
Creates a formatted .xlsx workbook:
  - Summary  : stat cards, top hype drops, breakdown tables, weekly outlook
  - Charts   : bar/pie charts for brand, hype, sale method, and price range
  - All Releases  : full sorted table
  - High Hype Alerts : HIGH + EXTREME only
"""

import os
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

NAVY        = "1F3864"
LIGHT_BLUE  = "C9D9EF"
WHITE       = "FFFFFF"
COL_LOW     = "00B050"   # green
COL_MED     = "FFC000"   # gold
COL_HIGH    = "FF0000"   # red
COL_EXTREME = "C00000"   # dark red

# Stat card accent colours
STAT_TOTAL  = "1F3864"   # navy
STAT_WEEK   = "C55A11"   # burnt orange
STAT_HYPE   = "C00000"   # dark red
STAT_PRICE  = "1F6B5C"   # teal

# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------

NAVY_FILL  = PatternFill(start_color=NAVY,       end_color=NAVY,       fill_type="solid")
BLUE_FILL  = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE,       end_color=WHITE,      fill_type="solid")

TITLE_FONT  = Font(name="Calibri", size=14, bold=True,  color=WHITE)
HEADER_FONT = Font(name="Calibri", size=11, bold=True,  color=WHITE)
DATA_FONT   = Font(name="Calibri", size=11,              color="000000")
MUTED_FONT  = Font(name="Calibri", size=9,  italic=True, color="888888")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

THIN_BORDER = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

HYPE_DISPLAY = {
    "EXTREME": ("EXTREME", COL_EXTREME),
    "HIGH":    ("HIGH",    COL_HIGH),
    "MEDIUM":  ("MED",     COL_MED),
    "LOW":     ("LOW",     COL_LOW),
}

SALE_METHOD_COLORS = {
    "SNKRS App":       "FF6B35",
    "Confirmed App":   "3B9EFF",
    "Raffle":          "FF9800",
    "Giveaway":        "43C96A",
    "Online":          "00BCD4",
    "Online + Retail": "888888",
    "In-Store":        "FFC107",
    "Retail":          "888888",
}

# ---------------------------------------------------------------------------
# Column definitions — All Releases / High Hype Alerts sheets
# (header, data-key, width, alignment, number-format)
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Date",        "release_date",  12, CENTER, "M/D/YYYY"),
    ("Retail",      "retail_price",  10, CENTER, "$#,##0"),
    ("Hype",        "hype_level",    11, CENTER, None),
    ("Sale Method", "sale_method",   17, CENTER, None),
    ("Brand",       "brand",         14, CENTER, None),
    ("Style",       "name",          46, CENTER, None),
]

NUM_COLS = len(COLUMNS)


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _cell(ws, row, col, value=None, font=None, fill=None,
          alignment=None, number_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font          = font
    if fill:          c.fill          = fill
    if alignment:     c.alignment     = alignment
    if number_format: c.number_format = number_format
    if border:        c.border        = border
    return c


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _font(size=11, bold=False, italic=False, color="000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def _merge_fill(ws, min_row, min_col, max_row, max_col, fill):
    """Fill every cell in a merged region (openpyxl only styles top-left)."""
    ws.merge_cells(start_row=min_row, start_column=min_col,
                   end_row=max_row,   end_column=max_col)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).fill = fill


# ---------------------------------------------------------------------------
# Shared title block used on release sheets
# ---------------------------------------------------------------------------

def _write_release_title(ws, label: str):
    last = get_column_letter(NUM_COLS)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    c1 = ws.cell(row=1, column=1, value=f"{label} Shoe Releases")
    c1.font = TITLE_FONT; c1.fill = NAVY_FILL; c1.alignment = CENTER
    ws.merge_cells(f"A1:{last}1")
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=1, column=col).fill = NAVY_FILL
    c2 = ws.cell(row=2, column=1, value="Calculated Hype Level")
    c2.font = _font(11, False, True, WHITE); c2.fill = NAVY_FILL; c2.alignment = CENTER
    ws.merge_cells(f"A2:{last}2")
    for col in range(2, NUM_COLS + 1):
        ws.cell(row=2, column=col).fill = NAVY_FILL


def _write_release_header(ws):
    ws.row_dimensions[3].height = 18
    for col_idx, (header, _, width, _, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = HEADER_FONT; c.fill = NAVY_FILL
        c.alignment = CENTER; c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_release_data(ws, sneakers):
    for i, sneaker in enumerate(sneakers):
        row_num = i + 4
        fill    = BLUE_FILL if i % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row_num].height = 15
        for col_idx, (_, key, _, align, num_fmt) in enumerate(COLUMNS, start=1):
            value = sneaker.get(key)
            if key == "release_date" and isinstance(value, datetime):
                value = value.date()
            elif key == "retail_price" and value is not None:
                value = float(value)
            elif key == "hype_level":
                pass
            elif value is None:
                value = ""
            c = ws.cell(row=row_num, column=col_idx, value=value)
            c.fill = fill; c.alignment = align; c.border = THIN_BORDER
            if num_fmt:
                c.number_format = num_fmt
            if key == "hype_level":
                display_text, color = HYPE_DISPLAY.get(value, (str(value), "000000"))
                c.value = display_text
                c.font  = _font(11, True, color=color)
            elif key == "sale_method":
                color = SALE_METHOD_COLORS.get(value, "888888")
                c.font = _font(11, True, color=color)
            else:
                c.font = DATA_FONT


def _create_release_sheet(wb, sneakers, title, sheet_name):
    ws = wb.create_sheet(sheet_name)
    _write_release_title(ws, title)
    _write_release_header(ws)
    if sneakers:
        _write_release_data(ws, sneakers)
    else:
        row = 4
        ws.cell(row=row, column=1, value="No releases found.").font = _font(11, False, True, "888888")
        ws.merge_cells(f"A{row}:{get_column_letter(NUM_COLS)}{row}")
        ws.cell(row=row, column=1).alignment = CENTER
    ws.freeze_panes = "A4"
    if sneakers:
        ws.auto_filter.ref = f"A3:{get_column_letter(NUM_COLS)}{len(sneakers) + 3}"


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
#
# Layout (8 columns: A–H)
#   A(30) B(12) C(11) D(10) E(11) F(9) G(12) H(16)
#
#  Rows 1–2   : navy title block
#  Rows 4–6   : 4 stat cards (2 cols each: A:B, C:D, E:F, G:H)
#  Row  8     : "Top Hype Releases" section header
#  Rows 9–...  : top-7 mini table  (Name A, Brand B, Date C, Days D, Hype E, Score F, Price G, Sale H)
#  Spacer row
#  Row  ...   : breakdown section headers (By Brand | By Hype | By Sale Method)
#  Rows ...   : three side-by-side breakdown tables (A:B | D:E | G:H)
#  Spacer row
#  Row  ...   : "Weekly Outlook" header
#  Rows ...   : 13-week release cadence
# ---------------------------------------------------------------------------

_SUM_COLS = 8        # A–H


def _sum_navy_row(ws, row, text, size=14, bold=True, italic=False, height=22):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(size, bold, italic, WHITE)
    c.fill = NAVY_FILL; c.alignment = CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_SUM_COLS)
    for col in range(2, _SUM_COLS + 1):
        ws.cell(row=row, column=col).fill = NAVY_FILL


def _stat_card(ws, rows, col_start, col_end, label, value, sub, color):
    """Write a 3-row stat card at given row range and column range."""
    label_row, val_row, sub_row = rows
    card_fill = _fill(color)

    # Label
    ws.row_dimensions[label_row].height = 16
    c = ws.cell(row=label_row, column=col_start, value=label)
    c.font = _font(9, True, color=WHITE); c.fill = card_fill; c.alignment = CENTER
    _merge_fill(ws, label_row, col_start, label_row, col_end, card_fill)
    ws.cell(row=label_row, column=col_start).font = _font(9, True, color=WHITE)
    ws.cell(row=label_row, column=col_start).alignment = CENTER
    ws.cell(row=label_row, column=col_start).value = label

    # Value
    ws.row_dimensions[val_row].height = 30
    c = ws.cell(row=val_row, column=col_start, value=value)
    c.font = _font(20, True, color=color); c.fill = WHITE_FILL; c.alignment = CENTER
    _merge_fill(ws, val_row, col_start, val_row, col_end, WHITE_FILL)
    ws.cell(row=val_row, column=col_start).font = _font(20, True, color=color)
    ws.cell(row=val_row, column=col_start).alignment = CENTER
    ws.cell(row=val_row, column=col_start).value = value

    # Sub-label
    ws.row_dimensions[sub_row].height = 13
    c = ws.cell(row=sub_row, column=col_start, value=sub)
    c.font = MUTED_FONT; c.fill = WHITE_FILL; c.alignment = CENTER
    _merge_fill(ws, sub_row, col_start, sub_row, col_end, WHITE_FILL)
    ws.cell(row=sub_row, column=col_start).font = MUTED_FONT
    ws.cell(row=sub_row, column=col_start).alignment = CENTER
    ws.cell(row=sub_row, column=col_start).value = sub

    # Outer border around the whole card
    card_border_color = "CCCCCC"
    thin = Side(style="thin", color=card_border_color)
    thick = Side(style="medium", color=color)
    for r in [label_row, val_row, sub_row]:
        for c_idx in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c_idx)
            left  = thick if c_idx == col_start  else Side(style="thin", color=card_border_color)
            right = thick if c_idx == col_end    else Side(style="thin", color=card_border_color)
            top   = thick if r == label_row else thin
            bot   = thick if r == sub_row   else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bot)


def _section_header(ws, row, text, height=16):
    """Dark-navy section divider spanning all summary columns."""
    ws.row_dimensions[row].height = height
    _sum_navy_row(ws, row, text, size=11, bold=True, height=height)


def _create_summary_sheet(wb, sneakers):
    ws = wb.create_sheet("Summary")

    # Column widths
    for col, w in zip("ABCDEFGH", [30, 12, 11, 10, 11, 9, 12, 16]):
        ws.column_dimensions[col].width = w

    # --- Computed stats ---
    total       = len(sneakers)
    this_week   = sum(1 for s in sneakers if (s.get("days_until_release") or 99) <= 7)
    high_ext    = sum(1 for s in sneakers if s.get("hype_level") in ("HIGH", "EXTREME"))
    prices      = [float(s["retail_price"]) for s in sneakers if s.get("retail_price")]
    avg_price   = sum(prices) / len(prices) if prices else 0

    # ---- TITLE BLOCK (rows 1–2) ----
    _sum_navy_row(ws, 1, "Sneaker Release Report — Summary", 14, True, height=26)
    _sum_navy_row(ws, 2,
        f"Generated {datetime.now(timezone.utc).strftime('%B %d, %Y  ·  %H:%M UTC')}",
        10, False, True, height=16)

    # Spacer row 3
    ws.row_dimensions[3].height = 6

    # ---- STAT CARDS (rows 4–6) ----
    card_rows = (4, 5, 6)
    _stat_card(ws, card_rows, 1, 2, "TOTAL RELEASES",    total,       "upcoming drops",  STAT_TOTAL)
    _stat_card(ws, card_rows, 3, 4, "DROPPING THIS WEEK", this_week,  "≤ 7 days out",    STAT_WEEK)
    _stat_card(ws, card_rows, 5, 6, "HIGH + EXTREME",    high_ext,    "hype releases",   STAT_HYPE)
    _stat_card(ws, card_rows, 7, 8, "AVG RETAIL PRICE",  f"${avg_price:.0f}", "across all drops", STAT_PRICE)

    # Spacer row 7
    ws.row_dimensions[7].height = 10

    # ---- TOP HYPE RELEASES (rows 8–16) ----
    _section_header(ws, 8, "Top Hype Releases")

    top_n     = sorted(sneakers, key=lambda s: s.get("hype_score", 0), reverse=True)[:7]
    mini_hdrs = ["Name", "Brand", "Date", "Days", "Hype", "Score", "Price", "Sale Method"]
    ws.row_dimensions[9].height = 15
    for ci, h in enumerate(mini_hdrs, 1):
        c = ws.cell(row=9, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    for i, s in enumerate(top_n):
        row = 10 + i
        fill = BLUE_FILL if i % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row].height = 15
        rd = s.get("release_date")
        if isinstance(rd, datetime):
            rd = rd.date()
        row_vals = [
            s.get("name", ""),
            s.get("brand", ""),
            rd,
            s.get("days_until_release", ""),
            s.get("hype_level", ""),
            s.get("hype_score", ""),
            s.get("retail_price"),
        ]
        # Sale method in col 8
        sale = s.get("sale_method") or ""

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=row, column=ci, value=val if val is not None else "")
            c.fill = fill; c.alignment = CENTER; c.border = THIN_BORDER
            if ci == 3 and val:
                c.number_format = "M/D/YYYY"
                c.font = DATA_FONT
            elif ci == 5:
                disp, color = HYPE_DISPLAY.get(val, (str(val), "000000"))
                c.value = disp
                c.font  = _font(11, True, color=color)
            elif ci == 7 and val:
                c.value = float(val)
                c.number_format = "$#,##0"
                c.font = DATA_FONT
            else:
                c.font = DATA_FONT

        # Sale method col 8
        c8 = ws.cell(row=row, column=8, value=sale)
        c8.fill = fill; c8.alignment = CENTER; c8.border = THIN_BORDER
        c8.font = _font(11, True, color=SALE_METHOD_COLORS.get(sale, "888888"))

    # Spacer
    spacer_r = 10 + len(top_n)
    ws.row_dimensions[spacer_r].height = 10

    # ---- BREAKDOWN TABLES (3 side-by-side) ----
    brk_title_r = spacer_r + 1
    _section_header(ws, brk_title_r, "Release Breakdown")

    # Sub-headers
    sub_r = brk_title_r + 1
    ws.row_dimensions[sub_r].height = 15
    for col, text in [(1, "Brand"), (4, "Hype Level"), (7, "Sale Method")]:
        c = ws.cell(row=sub_r, column=col, value=text)
        c.font = HEADER_FONT; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        ws.cell(row=sub_r, column=col + 1, value="Count").font = HEADER_FONT
        ws.cell(row=sub_r, column=col + 1).fill = NAVY_FILL
        ws.cell(row=sub_r, column=col + 1).alignment = CENTER
        ws.cell(row=sub_r, column=col + 1).border = THIN_BORDER

    # Brand breakdown (cols 1-2)
    brand_counts: dict[str, int] = {}
    for s in sneakers:
        brand_counts[s["brand"]] = brand_counts.get(s["brand"], 0) + 1

    # Hype breakdown (cols 4-5)
    hype_counts = {"EXTREME": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for s in sneakers:
        lvl = s.get("hype_level", "LOW")
        hype_counts[lvl] = hype_counts.get(lvl, 0) + 1

    # Sale method breakdown (cols 7-8)
    method_counts: dict[str, int] = {}
    for s in sneakers:
        m = s.get("sale_method") or "Online + Retail"
        method_counts[m] = method_counts.get(m, 0) + 1

    max_brk = max(
        len(brand_counts),
        len(hype_counts),
        len(method_counts),
    )

    brk_data_start = sub_r + 1
    brand_list = sorted(brand_counts, key=brand_counts.__getitem__, reverse=True)
    hype_list  = ["EXTREME", "HIGH", "MEDIUM", "LOW"]
    method_list = sorted(method_counts, key=method_counts.__getitem__, reverse=True)

    for i in range(max_brk):
        row = brk_data_start + i
        fill = BLUE_FILL if i % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row].height = 15

        # Brand
        if i < len(brand_list):
            b = brand_list[i]
            ws.cell(row=row, column=1, value=b).font = DATA_FONT
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2, value=brand_counts[b]).font = DATA_FONT
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = CENTER

        # Hype
        if i < len(hype_list):
            h = hype_list[i]
            disp, color = HYPE_DISPLAY.get(h, (h, "000000"))
            ws.cell(row=row, column=4, value=disp).font = _font(11, True, color=color)
            ws.cell(row=row, column=4).fill = fill
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=5, value=hype_counts[h]).font = DATA_FONT
            ws.cell(row=row, column=5).fill = fill
            ws.cell(row=row, column=5).border = THIN_BORDER
            ws.cell(row=row, column=5).alignment = CENTER

        # Sale method
        if i < len(method_list):
            m = method_list[i]
            color = SALE_METHOD_COLORS.get(m, "888888")
            ws.cell(row=row, column=7, value=m).font = _font(11, True, color=color)
            ws.cell(row=row, column=7).fill = fill
            ws.cell(row=row, column=7).border = THIN_BORDER
            ws.cell(row=row, column=7).alignment = LEFT
            ws.cell(row=row, column=8, value=method_counts[m]).font = DATA_FONT
            ws.cell(row=row, column=8).fill = fill
            ws.cell(row=row, column=8).border = THIN_BORDER
            ws.cell(row=row, column=8).alignment = CENTER

    # Spacer
    wk_sp = brk_data_start + max_brk
    ws.row_dimensions[wk_sp].height = 10

    # ---- WEEKLY OUTLOOK ----
    wk_hdr = wk_sp + 1
    _section_header(ws, wk_hdr, "Weekly Outlook (Next 13 Weeks)")

    wk_sub = wk_hdr + 1
    ws.row_dimensions[wk_sub].height = 15
    for col, text in [(1, "Week"), (2, "Dates"), (3, "Releases"), (4, "HIGH+EXTREME"), (5, "Avg Hype Score")]:
        c = ws.cell(row=wk_sub, column=col, value=text)
        c.font = HEADER_FONT; c.fill = NAVY_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    for w in range(13):
        wk_start = today + timedelta(weeks=w)
        wk_end   = wk_start + timedelta(days=6)
        row = wk_sub + 1 + w
        fill = BLUE_FILL if w % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row].height = 15

        wk_releases = [
            s for s in sneakers
            if isinstance(s.get("release_date"), datetime)
            and wk_start <= s["release_date"] <= wk_end
        ]
        wk_high = sum(1 for s in wk_releases if s.get("hype_level") in ("HIGH", "EXTREME"))
        scores  = [s.get("hype_score", 0) for s in wk_releases]
        avg_sc  = round(sum(scores) / len(scores), 1) if scores else ""

        week_label = f"Week {w + 1}"
        date_range = f"{wk_start.strftime('%-m/%-d')} – {wk_end.strftime('%-m/%-d')}"

        for col, val in [(1, week_label), (2, date_range), (3, len(wk_releases)),
                         (4, wk_high), (5, avg_sc)]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.border = THIN_BORDER; c.alignment = CENTER
            c.font = DATA_FONT if col != 4 or wk_high == 0 else _font(11, True, color=COL_HIGH)


# ---------------------------------------------------------------------------
# Charts sheet
# ---------------------------------------------------------------------------

def _create_charts_sheet(wb, sneakers):
    ws = wb.create_sheet("Charts")

    # --- Compute data ---
    brand_counts: dict[str, int] = {}
    for s in sneakers:
        brand_counts[s["brand"]] = brand_counts.get(s["brand"], 0) + 1

    hype_order  = ["EXTREME", "HIGH", "MEDIUM", "LOW"]
    hype_labels = ["EXTREME", "HIGH", "MED", "LOW"]
    hype_counts = {h: 0 for h in hype_order}
    for s in sneakers:
        lvl = s.get("hype_level", "LOW")
        if lvl in hype_counts:
            hype_counts[lvl] += 1

    method_counts: dict[str, int] = {}
    for s in sneakers:
        m = s.get("sale_method") or "Online + Retail"
        method_counts[m] = method_counts.get(m, 0) + 1

    price_buckets = ["< $100", "$100–$149", "$150–$199", "$200–$299", "$300+"]
    price_counts  = {b: 0 for b in price_buckets}
    for s in sneakers:
        p = s.get("retail_price")
        if p is None:
            continue
        p = float(p)
        if p < 100:
            price_counts["< $100"] += 1
        elif p < 150:
            price_counts["$100–$149"] += 1
        elif p < 200:
            price_counts["$150–$199"] += 1
        elif p < 300:
            price_counts["$200–$299"] += 1
        else:
            price_counts["$300+"] += 1

    # --- Write data tables ---

    # Col A:B — Brand
    ws.cell(row=1, column=1, value="Brand").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Releases").font = HEADER_FONT
    brand_list = sorted(brand_counts, key=brand_counts.__getitem__, reverse=True)
    for i, b in enumerate(brand_list, start=2):
        ws.cell(row=i, column=1, value=b)
        ws.cell(row=i, column=2, value=brand_counts[b])
    n_brand = len(brand_list)

    # Col D:E — Hype Level
    ws.cell(row=1, column=4, value="Hype Level").font = HEADER_FONT
    ws.cell(row=1, column=5, value="Releases").font = HEADER_FONT
    for i, (lvl, lbl) in enumerate(zip(hype_order, hype_labels), start=2):
        ws.cell(row=i, column=4, value=lbl)
        ws.cell(row=i, column=5, value=hype_counts[lvl])

    # Col G:H — Sale Method
    ws.cell(row=1, column=7, value="Sale Method").font = HEADER_FONT
    ws.cell(row=1, column=8, value="Releases").font = HEADER_FONT
    method_list = sorted(method_counts, key=method_counts.__getitem__, reverse=True)
    for i, m in enumerate(method_list, start=2):
        ws.cell(row=i, column=7, value=m)
        ws.cell(row=i, column=8, value=method_counts[m])
    n_method = len(method_list)

    # Col J:K — Price Range
    ws.cell(row=1, column=10, value="Price Range").font = HEADER_FONT
    ws.cell(row=1, column=11, value="Releases").font = HEADER_FONT
    for i, b in enumerate(price_buckets, start=2):
        ws.cell(row=i, column=10, value=b)
        ws.cell(row=i, column=11, value=price_counts[b])

    # --- Chart 1: Releases by Brand (vertical bar) ---
    c1 = BarChart()
    c1.type    = "col"
    c1.title   = "Releases by Brand"
    c1.y_axis.title = "Count"
    c1.style   = 10
    c1.width   = 18
    c1.height  = 12
    c1.legend  = None
    data1 = Reference(ws, min_col=2, min_row=1, max_row=n_brand + 1)
    cats1 = Reference(ws, min_col=1, min_row=2, max_row=n_brand + 1)
    c1.add_data(data1, titles_from_data=True)
    c1.set_categories(cats1)
    c1.series[0].graphicalProperties.solidFill = NAVY
    c1.series[0].graphicalProperties.line.solidFill = NAVY
    ws.add_chart(c1, "A" + str(n_brand + 4))

    # --- Chart 2: Hype Level Distribution (pie) ---
    c2 = PieChart()
    c2.title  = "Hype Level Distribution"
    c2.style  = 10
    c2.width  = 15
    c2.height = 12
    data2 = Reference(ws, min_col=5, min_row=1, max_row=5)
    cats2 = Reference(ws, min_col=4, min_row=2, max_row=5)
    c2.add_data(data2, titles_from_data=True)
    c2.set_categories(cats2)
    pie_colors = [COL_EXTREME, COL_HIGH, COL_MED, COL_LOW]
    for idx, color in enumerate(pie_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        c2.series[0].dPt.append(pt)
    ws.add_chart(c2, "J2")

    # --- Chart 3: Releases by Sale Method (horizontal bar) ---
    c3 = BarChart()
    c3.type    = "bar"   # horizontal
    c3.title   = "Releases by Sale Method"
    c3.x_axis.title = "Count"
    c3.style   = 10
    c3.width   = 18
    c3.height  = 12
    c3.legend  = None
    data3 = Reference(ws, min_col=8, min_row=1, max_row=n_method + 1)
    cats3 = Reference(ws, min_col=7, min_row=2, max_row=n_method + 1)
    c3.add_data(data3, titles_from_data=True)
    c3.set_categories(cats3)
    c3.series[0].graphicalProperties.solidFill = "3B9EFF"
    c3.series[0].graphicalProperties.line.solidFill = "3B9EFF"
    ws.add_chart(c3, "G" + str(n_method + 4))

    # --- Chart 4: Price Range Distribution (vertical bar) ---
    c4 = BarChart()
    c4.type    = "col"
    c4.title   = "Releases by Price Range"
    c4.y_axis.title = "Count"
    c4.style   = 10
    c4.width   = 15
    c4.height  = 12
    c4.legend  = None
    data4 = Reference(ws, min_col=11, min_row=1, max_row=6)
    cats4 = Reference(ws, min_col=10, min_row=2, max_row=6)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    c4.series[0].graphicalProperties.solidFill = COL_MED
    c4.series[0].graphicalProperties.line.solidFill = COL_MED
    ws.add_chart(c4, "J16")

    # Column widths for readability
    for col, w in zip("ABCDEFGHIJK", [14, 10, 3, 12, 10, 3, 18, 10, 3, 14, 10]):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _month_label(sneakers):
    dates = [s["release_date"] for s in sneakers if isinstance(s.get("release_date"), datetime)]
    if dates:
        return min(dates).strftime("%B %Y")
    return datetime.now().strftime("%B %Y")


def export_to_excel(sneakers: list[dict], output_path: str):
    """
    Export sneaker data to a formatted Excel workbook.

    Args:
        sneakers:    List of normalised sneaker dicts sorted by release date.
        output_path: Destination .xlsx file path.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    month = _month_label(sneakers)

    # Sheet order: Summary · Charts · All Releases · High Hype Alerts
    # Rename the default active sheet so _create_release_sheet can use create_sheet for all.
    wb.active.title = "_tmp"

    _create_summary_sheet(wb, sneakers)
    _create_charts_sheet(wb, sneakers)

    high_hype = [s for s in sneakers if s.get("hype_level") in ("HIGH", "EXTREME")]
    _create_release_sheet(wb, sneakers,  month, "All Releases")
    _create_release_sheet(wb, high_hype, month, "High Hype Alerts")

    # Remove the placeholder sheet
    del wb["_tmp"]

    wb.save(output_path)
